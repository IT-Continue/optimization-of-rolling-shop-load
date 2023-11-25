import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def main():
    pd.set_option('display.max_columns', None)

    file = pd.read_csv("answer_done.csv", index_col=0)

    num_cols = int(file.columns[-1].split('_')[1])

    file.insert(1, "Начальная температура", file["temp_1"].values)

    for i in range(1, num_cols + 1):
        file.drop(columns=f"temp_{ i }", inplace=True)

    file.to_excel("graph.xlsx", index=False)

    excel_file = load_workbook("graph.xlsx")
    sheet = excel_file.active

    colors = {
        "нагрев" : "FFA500",
        "подогрев" : "006400",
        "ковка" : "EE82EE",
        "прокат" : "00008B",
        "отжиг" : "FF0000",
        "nothing" : "808080"
    }

    for cell in sheet[1]:
        if isinstance(cell.value, str) and cell.value.startswith('operation_'):
            column = cell.column  # Column number

            for row in range(2, sheet.max_row + 1):  # Проходим по всем строкам, начиная со второй
                value = sheet.cell(row=row, column=column).value

                if value is not None:  # Cell is ot empty
                    color = colors.get(value.lower())  # Get color
                    if color:
                        sheet.cell(row=row, column=column).fill = PatternFill(start_color=color, end_color=color,
                                                                              fill_type="solid")

    excel_file.save("graph.xlsx")


if __name__ == "__main__":
    main()
